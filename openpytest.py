"""
This code automates recharges using the openpyxl library. The script takes data
from a compiled utility data excel sheet and inserts them into the recharge excel
sheets. The script takes the data from specific locations, so the data and the
recharge sheets need to stay in a consistent format, or data will be cell_input improperly.
See dataCompilation.xlsx for formating example.

Usage:
    Gets data from a data compilation excel sheet to store into Utility Recharges,
    Dining Utility Recharges, and Housing Recharges excel sheets. To change which
    to cell_input and output, go to individual methods and change the name of cell_input and
    where it is saved.
    Be sure to check data compilation for correct data before using the script.
TODO:
    1)Improve Documentation
    2)Implement more dynamic method of grabbing data so changing order of data
    won't matter anymore(Check row by name)
    3)Make it more convenient to change name of excel sheets being edited
"""
import re
from openpyxl import load_workbook
WORKBOOK2 = load_workbook('dataCompilation.xlsx')
def grab_values(data_sheet, cell_actual, cell_act, cell_exp):
    """
    Grabs values from data_sheet to be used to show the values being used for extrapolation

    Args:
    data_sheet(worksheet): The worksheet that contains the data
    cell_actual(str): The cell that contains actual recorded value from database.
    cell_act(str): The cell that contains actual # of data points counted from database.
    cell_exp(str): The cell that contains actual # of data points expected for that month

    Returns:
    Three variables containing the values of the requested cells.
    """
    actual = data_sheet[cell_actual].value
    num_actual = data_sheet[cell_act].value
    num_expected = data_sheet[cell_exp].value
    return actual, num_actual, num_expected

def calc_rates():
    """
    Calculates the rates from the data_sheet, and returns the values in equation
    form for double checking.

    Args:
    None

    Returns:
    Three variables containing the equations for kwh, water, and gas rates
    """
    #KWH Rate
    data_sheet = WORKBOOK2.worksheets[4]
    elcap_kwh = data_sheet['B1'].value
    elcap_price = data_sheet['B2'].value
    wilson_kwh = data_sheet['B3'].value
    wilson_price = data_sheet['B4'].value
    sunpower_kwh = data_sheet['B5'].value
    sunpower_price = data_sheet['B6'].value
    ucop = data_sheet['B7'].value
    kwh_rate = '=('+repr(elcap_price)+'+'+repr(wilson_price)+'+'+repr(sunpower_price)+'+'+repr(ucop)
    kwh_rate = kwh_rate+')/('+repr(elcap_kwh)+'+'+repr(wilson_kwh)+'+'+repr(sunpower_kwh)+')'
    #Water Rate
    h2o_price = data_sheet['B8'].value
    h2o_hcf = data_sheet['B9'].value
    water_rate = '='+repr(h2o_price)+'/(' +repr(h2o_hcf)+'*(748/1000))'
    #Gas Rate
    gas_price = data_sheet['B10'].value
    gas_therms = data_sheet['B11'].value
    gas_rate = '='+repr(gas_price)+'/' +repr(gas_therms)
    return kwh_rate, water_rate, gas_rate

def get_columns():
    """
    Finds which column to cell_input for all recharges. Uses a while loop so
    that it can handle invalid inputs.

    Args:
    None

    Returns:
    int containing column number to cell_input all data for recharges
    """
    #If statements to handle months that are out of order. Columns C(july)-N(june)
    while True:
        month = int(input("Input Month #(1-12): "))
        if month > 12 or month < 1:
            error = 'Invalid Month, please cell_input a month from 1-12'
            str(error)
        elif month >= 7:
            input_column = 3+(month-7)
            break
        else:
            input_column = 3 + month + 5
            break
    return input_column

def fill_utility(utility_name, input_column, kwh_rate, water_rate, gas_rate):
    """
    Fills in data for Gall R&W, Facilities, and the Library.

    Args:
    input_column(int): Determines which column(month) is being filled in.
    kwh_rate(str): The kwhrate to be filled in the recharge
    water_rate(str): The kwhrate to be filled in the recharge
    gas_rate(str): The kwhrate to be filled in the recharge

    Returns:
    None
    """
    workbook = load_workbook(utility_name)
    #Editing Gall R&W Summary
    edit_sheet = workbook["Gall R&W Utility Summary"]
    #Electricity Input
    data_sheet = WORKBOOK2.worksheets[0]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C2', 'E2', 'F2')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=7, column=input_column).value = cell_input
    edit_sheet.cell(row=8, column=input_column).value = kwh_rate

    #Gas Input
    data_sheet = WORKBOOK2.worksheets[1]
    reading = data_sheet['C11'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=12, column=input_column).value = reading
    edit_sheet.cell(row=14, column=input_column).value = gas_rate

    #CHW cell_input
    data_sheet = WORKBOOK2.worksheets[2]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C2', 'E2', 'F2')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=30, column=input_column).value = cell_input

    #Water Input
    data_sheet = WORKBOOK2.worksheets[3]
    edit_sheet.cell(row=20, column=input_column).value = water_rate
    edit_sheet.cell(row=26, column=input_column).value = water_rate
    reading = data_sheet['C39'].value+data_sheet['C40'].value-data_sheet['D39'].value-data_sheet['D40'].value
    edit_sheet.cell(row=19, column=input_column).value = reading
    reading1 = data_sheet['C61'].value-data_sheet['D61'].value
    reading2 = data_sheet['C62'].value-data_sheet['D62'].value
    edit_sheet.cell(row=24, column=input_column).value = reading1
    edit_sheet.cell(row=25, column=input_column).value = reading2

    #Editing Facilities Summary
    edit_sheet = workbook["Facilities Utility Summary"]
    #Electricity Input
    data_sheet = WORKBOOK2.worksheets[0]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C3', 'E3', 'F3')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=6, column=input_column).value = cell_input
    edit_sheet.cell(row=7, column=input_column).value = kwh_rate

    #Water Input
    data_sheet = WORKBOOK2.worksheets[3]
    edit_sheet.cell(row=17, column=input_column).value = water_rate
    reading = data_sheet['C26'].value+data_sheet['C27'].value-data_sheet['D26'].value-data_sheet['D27'].value
    edit_sheet.cell(row=16, column=input_column).value = reading

    #CHW cell_input
    data_sheet = WORKBOOK2.worksheets[2]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C3', 'E3', 'F3')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=21, column=input_column).value = cell_input

    #Editing Library Utility Summary
    edit_sheet = workbook["Library Utility Summary"]
    #Electricity Input
    data_sheet = WORKBOOK2.worksheets[0]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C4', 'E4', 'F4')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=6, column=input_column).value = cell_input
    edit_sheet.cell(row=7, column=input_column).value = kwh_rate

    #Gas Input
    data_sheet = WORKBOOK2.worksheets[1]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C15', 'E15', 'F15')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=11, column=input_column).value = cell_input
    edit_sheet.cell(row=12, column=input_column).value = gas_rate

    #CHW cell_input
    data_sheet = WORKBOOK2.worksheets[2]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C4', 'E4', 'F4')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=21, column=input_column).value = cell_input

    #Water Input
    data_sheet = WORKBOOK2.worksheets[3]
    edit_sheet.cell(row=17, column=input_column).value = water_rate
    reading = data_sheet['C34'].value+data_sheet['C35'].value-data_sheet['D34'].value-data_sheet['D35'].value
    edit_sheet.cell(row=16, column=input_column).value = reading

    #Save Changes to Utilities excel sheet
    workbook.save(utility_name)

def fill_dining(dining_name, input_column, kwh_rate, water_rate):
    """
    Fills in data for Dining, Laundry

    Args:
    input_column(int): Determines which column(month) is being filled in.
    kwh_rate(str): The kwhrate to be filled in the recharge
    water_rate(str): The kwhrate to be filled in the recharge
    gas_rate(str): The kwhrate to be filled in the recharge

    Returns:
    None
    """
    #Dining Utility Summary Fill out
    workbook = load_workbook(dining_name)
    #Electricity Input
    edit_sheet = workbook["Electricity"]
    data_sheet = WORKBOOK2.worksheets[0]

    #Dining Commons KWH
    edit_sheet.cell(row=5, column=input_column).value = kwh_rate
    actual, num_actual, num_expected = grab_values(data_sheet, 'C6', 'E6', 'F6')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=8, column=input_column).value = cell_input

    #Dining Expansion KWH
    actual, num_actual, num_expected = grab_values(data_sheet, 'C5', 'E5', 'F5')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=11, column=input_column).value = cell_input

    #Gas Input
    edit_sheet = workbook["Gas"]
    data_sheet = WORKBOOK2.worksheets[1]

    #Dining Boiler
    reading = data_sheet['C6'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=7, column=input_column).value = reading

    #Laundry
    reading = data_sheet['C7'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=11, column=input_column).value = reading

    #CHW Input
    edit_sheet = workbook["Chilled Water"]
    data_sheet = WORKBOOK2.worksheets[2]
    actual, num_actual, num_expected = grab_values(data_sheet, 'C5', 'E5', 'F5')
    cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
    edit_sheet.cell(row=7, column=input_column).value = cell_input

    #Water Input
    edit_sheet = workbook["Water"]
    data_sheet = WORKBOOK2.worksheets[3]
    edit_sheet.cell(row=5, column=input_column).value = water_rate
    reading = data_sheet['C34'].value+data_sheet['C35'].value-data_sheet['D34'].value-data_sheet['D35'].value
    edit_sheet.cell(row=8, column=input_column).value = reading

    #Saving Changes to Dining
    workbook.save(dining_name)

def fill_housing(housing_name, input_column, kwh_rate, water_rate, gas_rate):
    """
    Fills in recharges for Sierra Terraces, Valley Terraces, and Summits

    Args:
    input_column(int): Determines which column(month) is being filled in.
    kwh_rate(str): The kwhrate to be filled in the recharge
    water_rate(str): The kwhrate to be filled in the recharge
    gas_rate(str): The kwhrate to be filled in the recharge

    Returns:
    None
    """
    #Housing Utility Summary Fill out
    workbook = load_workbook(housing_name)
    #Electricity Input
    edit_sheet = workbook["Electricity"]
    data_sheet = WORKBOOK2.worksheets[0]
    edit_sheet.cell(row=5, column=input_column).value = kwh_rate
    for i in range(1, 6):
        actual = data_sheet.cell(row=i+6, column=3).value
        num_actual = data_sheet.cell(row=i+6, column=5).value
        num_expected = data_sheet.cell(row=i+6, column=6).value
        cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
        edit_sheet.cell(row=3+i*3, column=input_column).value = cell_input
        actual = data_sheet.cell(row=12, column=3).value
        edit_sheet.cell(row=21, column=input_column).value = actual
    for i in range(7, 15):
        actual = data_sheet.cell(row=i+6, column=3).value
        num_actual = data_sheet.cell(row=i+6, column=5).value
        num_expected = data_sheet.cell(row=i+6, column=6).value
        cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
        edit_sheet.cell(row=4+i*3, column=input_column).value = cell_input

    #Gas Input
    edit_sheet = workbook["Gas"]
    data_sheet = WORKBOOK2.worksheets[1]
    edit_sheet.cell(row=5, column=input_column).value = gas_rate

    #Dining
    reading = data_sheet['C6'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=7, column=input_column).value = reading

    #Laundry
    reading = data_sheet['C7'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=11, column=input_column).value = reading

    #Sierra Terraces
    reading = data_sheet['C10'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=15, column=input_column).value = reading

    #Tenaya
    reading = data_sheet['C4'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=20, column=input_column).value = reading

    #Cathedral
    reading = data_sheet['C3'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=25, column=input_column).value = reading

    #Half Dome
    reading = data_sheet['C5'].value
    reading = re.sub('cuft.', '', reading)
    edit_sheet.cell(row=30, column=input_column).value = reading

    #Water Input
    data_sheet = WORKBOOK2.worksheets[3]
    edit_sheet = workbook["Water"]
    edit_sheet.cell(row=5, column=input_column).value = water_rate

    #Terrace Center Commons
    reading = data_sheet['C68'].value+data_sheet['C69'].value-data_sheet['D68'].value-data_sheet['D69'].value
    edit_sheet.cell(row=8, column=input_column).value = reading

    #Kern
    reading = data_sheet['C32'].value+data_sheet['C33'].value-data_sheet['D32'].value-data_sheet['D33'].value
    edit_sheet.cell(row=12, column=input_column).value = reading

    #Tulare
    reading = data_sheet['C70'].value+data_sheet['C71'].value-data_sheet['D70'].value-data_sheet['D71'].value
    edit_sheet.cell(row=16, column=input_column).value = reading

    #Madera+Fresno+Stan+Kings
    reading = data_sheet['C57'].value+data_sheet['C58'].value-data_sheet['D57'].value-data_sheet['D58'].value
    reading = reading/4
    for i in range(1, 5):
        edit_sheet.cell(row=15+i*4, column=input_column).value = reading

    #San Joaquin
    reading = data_sheet['C46'].value+data_sheet['C47'].value-data_sheet['D46'].value-data_sheet['D47'].value
    edit_sheet.cell(row=35, column=input_column).value = reading

    #Merced+Calaveras
    reading = data_sheet['C4'].value+data_sheet['C5'].value-data_sheet['D4'].value-data_sheet['D5'].value
    reading = reading/2
    edit_sheet.cell(row=38, column=input_column).value = reading
    edit_sheet.cell(row=41, column=input_column).value = reading

    #Sierra Terraces
    reading = data_sheet['C72'].value+data_sheet['C73'].value-data_sheet['D72'].value-data_sheet['D73'].value
    edit_sheet.cell(row=44, column=input_column).value = reading

    #Tenaya
    reading = data_sheet['C66'].value+data_sheet['C67'].value-data_sheet['D66'].value-data_sheet['D67'].value
    edit_sheet.cell(row=47, column=input_column).value = reading

    #Cathedral
    reading = data_sheet['C10'].value+data_sheet['C11'].value-data_sheet['D10'].value-data_sheet['D11'].value
    edit_sheet.cell(row=50, column=input_column).value = reading

    #Half Dome
    reading = data_sheet['C28'].value+data_sheet['C29'].value-data_sheet['D28'].value-data_sheet['D29'].value
    edit_sheet.cell(row=53, column=input_column).value = reading

    #CHW cell_input
    edit_sheet = workbook["Chilled Water"]
    data_sheet = WORKBOOK2.worksheets[2]
    for i in range(1, 15):
        actual = data_sheet.cell(row=i+5, column=3).value
        num_actual = data_sheet.cell(row=i+5, column=5).value
        num_expected = data_sheet.cell(row=i+5, column=6).value
        cell_input = '='+repr(actual)+'/('+repr(num_actual)+'/'+repr(num_expected)+')'
        edit_sheet.cell(row=3+i*3, column=input_column).value = cell_input
    #Saving Changes to Housing excel sheet
    workbook.save(housing_name)

def main():
    """
    Runs the script methods. Besides getting the rates and input column, the
    fill method order doesn't matter. Change the names of the excel sheets to
    be edited here, in single quotes full filename(e.g 'test1.xlsx')
    """
    utility_name = 'test1.xlsx'
    dining_name = 'test2.xlsx'
    housing_name = 'test3.xlsx'
    input_column = get_columns()
    kwh_rate, water_rate, gas_rate = calc_rates()
    fill_utility(utility_name, input_column, kwh_rate, water_rate, gas_rate)
    fill_dining(dining_name, input_column, kwh_rate, water_rate)
    fill_housing(housing_name, input_column, kwh_rate, water_rate, gas_rate)

if __name__ == "__main__":
    main()
